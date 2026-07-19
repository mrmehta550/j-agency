"""
home/views.py

Production-quality contact & booking form implementation.

Key design decisions
--------------------
* All raw request.POST reads replaced with Django Form validation.
* logging.getLogger(__name__) replaces every print() call so errors
  appear in WSGI logs / Sentry in production.
* transaction.atomic() + select_for_update() prevents race-condition
  duplicate submissions at the database level for BOTH forms.
* Email side-effects run AFTER the transaction commits so a DB rollback
  never triggers a sent email.
* Excel writes are guarded by a module-level threading.Lock so two
  concurrent Gunicorn threads cannot corrupt the file simultaneously.
* PRG (Post/Redirect/Get) is applied on every POST path — success,
  duplicate, and validation failure — so the browser back-button never
  re-submits a form.
* Old ContactSubmissionLock rows are cleaned up after each successful
  contact submission to keep the table small.
* Booking submissions are fingerprinted on email+phone+service+date+time
  so a user cannot accidentally book the same slot twice.
"""

import hashlib
import logging
import os
import threading
from datetime import timedelta

from django.conf import settings
from django.contrib import messages
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from blog.models import Blog, Category

from .forms import BookingForm, ContactForm
from .models import Contact, ContactSubmissionLock


# ──────────────────────────────────────────────────────────────
# Module-level constants & singletons
# ──────────────────────────────────────────────────────────────

logger = logging.getLogger(__name__)

# How long a fingerprint blocks re-submission (2-minute window).
CONTACT_DUPLICATE_WINDOW = timedelta(minutes=2)

# Booking duplicate window — same email+phone+service+date+time
# within 5 minutes is treated as an accidental double-click.
BOOKING_DUPLICATE_WINDOW = timedelta(minutes=5)

# How long before an old submission lock is eligible for cleanup.
CONTACT_LOCK_MAX_AGE = timedelta(hours=24)

# Protects Excel file writes across threads in the same process.
# For multi-process deployments (multiple Gunicorn workers) consider
# using a file-system lock (fcntl / portalocker) instead.
_excel_write_lock = threading.Lock()


# ──────────────────────────────────────────────────────────────
# Simple page views
# ──────────────────────────────────────────────────────────────

def home(request):
    return render(request, "home/index.html")


def price(request):
    return render(request, "home/pricing.html")


def about(request):
    return render(request, "home/about.html")


def tech(request):
    return render(request, "home/tech.html")


def privacy(request):
    return render(request, "home/privacy.html")


def term(request):
    return render(request, "home/terms.html")


def cloud(request):
    return render(request, "home/cloud.html")


# ──────────────────────────────────────────────────────────────
# Contact form helpers
# ──────────────────────────────────────────────────────────────

def _build_contact_fingerprint(email: str, phone: str, service: str, message: str) -> str:
    """
    Build a SHA-256 fingerprint from the submission's key fields.

    The fingerprint is used as a deduplication key in ContactSubmissionLock.
    Email and phone are already normalised (lowercase / digits-only) by the
    form's clean methods before this function is called.

    full_name is intentionally excluded: two different people sharing the
    same email, phone, service, and message body are unlikely in practice,
    and including the name would let bad actors bypass the lock by changing
    a single character.
    """
    raw = "|".join([email.lower(), phone, service, message])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _save_contact_to_excel(contact: Contact) -> None:
    """
    Append a contact record to the shared contacts.xlsx workbook.

    Thread-safety: guarded by _excel_write_lock so concurrent requests
    cannot interleave their reads and writes to the file.
    """
    with _excel_write_lock:
        if not os.path.exists(settings.MEDIA_ROOT):
            os.makedirs(settings.MEDIA_ROOT)

        excel_file = os.path.join(settings.MEDIA_ROOT, "contacts.xlsx")

        # Load the existing workbook or create a fresh one if it is
        # missing, empty, or corrupt.
        if os.path.exists(excel_file):
            try:
                workbook = load_workbook(excel_file)
                worksheet = workbook.active
            except (OSError, InvalidFileException, PermissionError, EOFError):
                logger.warning(
                    "contacts.xlsx could not be opened — creating a new workbook.",
                    exc_info=True,
                )
                workbook = Workbook()
                worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active

        # Write header row on a brand-new sheet.
        if worksheet.max_row == 1 and worksheet["A1"].value is None:
            worksheet.append([
                "Name", "Email", "Phone", "Company",
                "Service", "Budget", "Message", "Created At",
            ])

        worksheet.append([
            contact.full_name,
            contact.email,
            contact.phone,
            contact.company,
            contact.service,
            contact.budget,
            contact.message,
            contact.created_at.strftime("%d-%m-%Y %H:%M:%S"),
        ])

        workbook.save(excel_file)
        workbook.close()


def _send_contact_emails(contact: Contact) -> None:
    """
    Send a notification email to the admin and a confirmation to the client.

    Called OUTSIDE the atomic block so a transient email error never
    rolls back the already-committed database record.
    """
    # ── Admin notification ──────────────────────────────────────
    admin_subject = f"🚀 New Contact Request - {contact.full_name}"
    admin_message = f"""\
🚀 NEW CONTACT REQUEST

━━━━━━━━━━━━━━━━━━━━━━━━━━

👤 Name:
{contact.full_name}

📧 Email:
{contact.email}

📱 Phone:
{contact.phone if contact.phone else "Not Provided"}

🏢 Company:
{contact.company if contact.company else "Not Provided"}

💼 Service:
{contact.service}

💰 Budget:
{contact.budget}

📝 Project Details

{contact.message}

━━━━━━━━━━━━━━━━━━━━━━━━━━

Submitted At:
{contact.created_at.strftime("%d-%m-%Y %H:%M")}
"""

    send_mail(
        subject=admin_subject,
        message=admin_message,
        from_email=settings.EMAIL_HOST_USER,
        recipient_list=[settings.EMAIL_HOST_USER],
        fail_silently=False,
    )

    # ── Client confirmation ─────────────────────────────────────
    client_subject = "✅ We've Received Your Request | Magency"
    client_message = f"""\
Hi {contact.full_name},

Thank you for contacting Magency.

We have successfully received your project request.

━━━━━━━━━━━━━━━━━━━━━━━━━━

YOUR SUBMISSION

👤 Name:
{contact.full_name}

📧 Email:
{contact.email}

📱 Phone:
{contact.phone}

🏢 Company:
{contact.company if contact.company else "Not Provided"}

💼 Service:
{contact.service}

💰 Budget:
{contact.budget}

📝 Project Details

{contact.message}

━━━━━━━━━━━━━━━━━━━━━━━━━━

Our team will carefully review your requirements.

You can expect a response within the next 24 hours.

If your project is urgent, simply reply to this email.

Thank you for choosing Magency.

Best Regards,

Magency Team

🌐 https://magencyinfo.com
📧 magency550@gmail.com
"""

    send_mail(
        subject=client_subject,
        message=client_message,
        from_email=settings.EMAIL_HOST_USER,
        recipient_list=[contact.email],
        fail_silently=False,
    )


def _cleanup_old_submission_locks() -> None:
    """
    Delete ContactSubmissionLock rows that are older than CONTACT_LOCK_MAX_AGE.

    Called after every successful submission.  Failures are logged but do not
    affect the user-facing response — the cleanup is purely housekeeping.
    """
    try:
        cutoff = timezone.now() - CONTACT_LOCK_MAX_AGE
        deleted_count, _ = ContactSubmissionLock.objects.filter(
            last_submitted_at__lt=cutoff
        ).delete()
        if deleted_count:
            logger.info("Cleaned up %d expired ContactSubmissionLock rows.", deleted_count)
    except Exception:
        logger.exception("Failed to clean up old ContactSubmissionLock rows.")


# ──────────────────────────────────────────────────────────────
# Contact view  (PRG pattern)
# ──────────────────────────────────────────────────────────────

def contact(request):
    """
    GET  → render the blank contact form.
    POST → validate → deduplicate → save → redirect (PRG).

    PRG (Post / Redirect / Get) is applied on every POST code path so
    the browser back-button or a page refresh never re-submits the form.
    """
    if request.method != "POST":
        return render(request, "home/contact.html")

    form = ContactForm(request.POST)

    if not form.is_valid():
        # Collect all field errors into a single readable message.
        error_messages = []
        for field, errors in form.errors.items():
            label = form.fields[field].label or field.replace("_", " ").title()
            for error in errors:
                error_messages.append(f"{label}: {error}")

        messages.error(
            request,
            " | ".join(error_messages) if error_messages else "Please fill all required fields correctly.",
        )
        return redirect("contact")   # PRG — keeps the URL clean

    # ── Extract validated, normalised data ──────────────────────
    cd = form.cleaned_data
    full_name = cd["full_name"]
    email     = cd["email"]       # already lowercased by clean_email()
    phone     = cd["phone"]       # already normalised by clean_phone()
    company   = cd.get("company", "")
    service   = cd["service"]
    budget    = cd["budget"]
    message   = cd["message"]

    submission_fingerprint = _build_contact_fingerprint(
        email=email,
        phone=phone,
        service=service,
        message=message,
    )

    # ── Atomic duplicate-check + DB write ───────────────────────
    # select_for_update() locks the matching row for the duration of
    # the transaction, so two concurrent requests with the same
    # fingerprint cannot both pass the duplicate check simultaneously.
    try:
        with transaction.atomic():
            duplicate_cutoff = timezone.now() - CONTACT_DUPLICATE_WINDOW

            is_duplicate = (
                ContactSubmissionLock.objects
                .select_for_update()
                .filter(
                    submission_fingerprint=submission_fingerprint,
                    last_submitted_at__gte=duplicate_cutoff,
                )
                .exists()
            )

            if is_duplicate:
                messages.warning(
                    request,
                    "You have already submitted this request recently. "
                    "Please wait a couple of minutes before trying again.",
                )
                return redirect("contact")   # PRG

            # Upsert the lock record — creates it on first submission,
            # updates the timestamp on subsequent (allowed) ones.
            ContactSubmissionLock.objects.update_or_create(
                submission_fingerprint=submission_fingerprint,
                defaults={"last_submitted_at": timezone.now()},
            )

            # Persist the contact record inside the same transaction so
            # the lock and the contact are always written together.
            contact_record = Contact.objects.create(
                full_name=full_name,
                email=email,
                phone=phone,
                company=company,
                service=service,
                budget=budget,
                message=message,
                project_details=message,   # kept for compatibility
            )

    except IntegrityError:
        # The unique constraint on submission_fingerprint was hit by a
        # concurrent request that snuck past the select_for_update check
        # (e.g., first submission on an empty table).
        logger.warning(
            "IntegrityError on ContactSubmissionLock for fingerprint %s — "
            "treating as duplicate submission.",
            submission_fingerprint,
        )
        messages.warning(
            request,
            "You have already submitted this request recently. "
            "Please wait a couple of minutes before trying again.",
        )
        return redirect("contact")   # PRG

    except Exception:
        logger.exception("Unexpected error while saving contact form submission.")
        messages.error(
            request,
            "Something went wrong while submitting your request. Please try again.",
        )
        return redirect("contact")   # PRG

    # ── Post-commit side-effects ─────────────────────────────────
    # These run AFTER the transaction commits.  If they fail the DB
    # record is already safely saved; we log the error but do not
    # surface it to the user (the submission succeeded).

    try:
        _save_contact_to_excel(contact_record)
    except Exception:
        logger.exception(
            "Failed to write contact #%d to Excel.", contact_record.pk
        )

    try:
        _send_contact_emails(contact_record)
    except Exception:
        logger.exception(
            "Failed to send notification emails for contact #%d.", contact_record.pk
        )

    # Housekeeping — best-effort, never blocks the response.
    _cleanup_old_submission_locks()

    messages.success(
        request,
        "🎉 Thank you! Your request has been submitted successfully. "
        "Our team will contact you within 24 hours.",
    )
    return redirect("contact")   # PRG — final redirect on success


# ───────────────────────────────────────────────────────────────
# Booking helpers
# ───────────────────────────────────────────────────────────────

def _build_booking_fingerprint(
    email: str,
    phone: str,
    service: str,
    preferred_date: str,
    preferred_time: str,
) -> str:
    """
    Build a SHA-256 fingerprint for a booking submission.

    Keyed on email + phone + service + preferred_date + preferred_time so
    that two requests for the exact same consultation slot by the same
    person are blocked, while different slots from the same person are
    always allowed through.

    preferred_date and preferred_time are converted to strings before
    hashing so that None values (optional fields) hash consistently.
    """
    raw = "|".join([
        email.lower(),
        phone,
        service,
        str(preferred_date),
        str(preferred_time),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _send_booking_emails(booking: Contact) -> None:
    """
    Send a booking-confirmation email to the admin and a receipt to the client.

    Called OUTSIDE the atomic block so a transient SMTP error never rolls
    back the already-committed booking record.
    """
    preferred_date_str = (
        booking.preferred_date.strftime("%d %B %Y")
        if booking.preferred_date
        else "Not specified"
    )
    preferred_time_str = (
        booking.preferred_time.strftime("%I:%M %p")
        if booking.preferred_time
        else "Not specified"
    )

    # ── Admin notification ─────────────────────────────────
    admin_subject = f"📅 New Booking Request - {booking.full_name}"
    admin_message = f"""\
📅 NEW BOOKING REQUEST

━━━━━━━━━━━━━━━━━━━━━━━━━━

👤 Name:
{booking.full_name}

📧 Email:
{booking.email}

📱 Phone:
{booking.phone if booking.phone else "Not Provided"}

🏢 Company:
{booking.company if booking.company else "Not Provided"}

💼 Service:
{booking.service}

💰 Budget:
{booking.budget}

🗓️ Preferred Date:
{preferred_date_str}

⏰ Preferred Time:
{preferred_time_str}

📝 Project Details

{booking.project_details}

━━━━━━━━━━━━━━━━━━━━━━━━━━

Booked At:
{booking.created_at.strftime("%d-%m-%Y %H:%M")}
"""

    send_mail(
        subject=admin_subject,
        message=admin_message,
        from_email=settings.EMAIL_HOST_USER,
        recipient_list=[settings.EMAIL_HOST_USER],
        fail_silently=False,
    )

    # ── Client confirmation ─────────────────────────
    client_subject = "✅ Your Consultation is Booked | Magency"
    client_message = f"""\
Hi {booking.full_name},

Thank you for booking a consultation with Magency.

Our team will reach out within 30 minutes to confirm your session.

━━━━━━━━━━━━━━━━━━━━━━━━━━

YOUR BOOKING DETAILS

👤 Name:
{booking.full_name}

📧 Email:
{booking.email}

📱 Phone:
{booking.phone}

🏢 Company:
{booking.company if booking.company else "Not Provided"}

💼 Service:
{booking.service}

💰 Budget:
{booking.budget}

🗓️ Preferred Date:
{preferred_date_str}

⏰ Preferred Time:
{preferred_time_str}

📝 Project Details

{booking.project_details}

━━━━━━━━━━━━━━━━━━━━━━━━━━

If you need to reschedule or have any questions,
simply reply to this email.

Thank you for choosing Magency.

Best Regards,

Magency Team

🌐 https://magencyinfo.com
📧 magency550@gmail.com
"""

    send_mail(
        subject=client_subject,
        message=client_message,
        from_email=settings.EMAIL_HOST_USER,
        recipient_list=[booking.email],
        fail_silently=False,
    )


# ──────────────────────────────────────────────────────────────
# Booking view  (PRG + fingerprint deduplication + atomic write)
# ──────────────────────────────────────────────────────────────

def booking(request):
    """
    GET  → render blank booking form.
    POST → validate via BookingForm → fingerprint → atomic duplicate-check
          → save → send emails → redirect (PRG).

    Duplicate detection
    -------------------
    A SHA-256 fingerprint of email + phone + service + preferred_date +
    preferred_time is checked against ContactSubmissionLock inside a
    select_for_update() transaction.  Two concurrent requests with the
    same fingerprint cannot both pass simultaneously; the second will
    either be caught by the lock query or hit the unique IntegrityError.

    PRG on every code path
    ----------------------
    Validation failure, duplicate, DB error, and success all end with
    redirect() so the browser back-button never re-POSTs.
    """
    if request.method != "POST":
        return render(request, "home/booking.html")

    form = BookingForm(request.POST)

    if not form.is_valid():
        # Collect all field-level errors into one readable flash message.
        error_messages = []
        for field, errors in form.errors.items():
            label = form.fields[field].label or field.replace("_", " ").title()
            for error in errors:
                error_messages.append(f"{label}: {error}")

        messages.error(
            request,
            " | ".join(error_messages) if error_messages else "Please correct the errors below.",
        )
        return redirect("booking")   # PRG

    # ── Extract validated, normalised data ──────────────────────
    cd = form.cleaned_data
    email          = cd["email"]                  # lowercased by clean_email()
    phone          = cd["phone"]                  # normalised by clean_phone()
    service        = cd["service"]
    preferred_date = cd.get("preferred_date")     # optional DateField
    preferred_time = cd.get("preferred_time")     # optional TimeField

    booking_fingerprint = _build_booking_fingerprint(
        email=email,
        phone=phone,
        service=service,
        preferred_date=preferred_date,
        preferred_time=preferred_time,
    )

    # ── Atomic duplicate-check + DB write ───────────────────────
    # select_for_update() row-locks the matching ContactSubmissionLock
    # record for the duration of the transaction.  A second concurrent
    # request with the same fingerprint blocks at the lock and, after
    # the first request commits, finds the record and is redirected.
    try:
        with transaction.atomic():
            duplicate_cutoff = timezone.now() - BOOKING_DUPLICATE_WINDOW

            is_duplicate = (
                ContactSubmissionLock.objects
                .select_for_update()
                .filter(
                    submission_fingerprint=booking_fingerprint,
                    last_submitted_at__gte=duplicate_cutoff,
                )
                .exists()
            )

            if is_duplicate:
                messages.warning(
                    request,
                    "It looks like you already submitted this booking recently. "
                    "Please wait a few minutes before trying again, or choose "
                    "a different date or time slot.",
                )
                return redirect("booking")   # PRG

            # Upsert the lock record — created on first submission,
            # timestamp refreshed on any later allowed re-submission.
            ContactSubmissionLock.objects.update_or_create(
                submission_fingerprint=booking_fingerprint,
                defaults={"last_submitted_at": timezone.now()},
            )

            # Save the booking inside the same transaction so the lock
            # and the record are always committed or rolled back together.
            booking_record = form.save()

    except IntegrityError:
        # The unique constraint on submission_fingerprint was hit by a
        # concurrent request that raced past the select_for_update check.
        logger.warning(
            "IntegrityError on ContactSubmissionLock for booking fingerprint %s — "
            "treating as duplicate submission.",
            booking_fingerprint,
        )
        messages.warning(
            request,
            "It looks like you already submitted this booking recently. "
            "Please wait a few minutes before trying again.",
        )
        return redirect("booking")   # PRG

    except Exception:
        logger.exception("Unexpected error while saving booking form submission.")
        messages.error(
            request,
            "Something went wrong while submitting your booking. Please try again.",
        )
        return redirect("booking")   # PRG

    # ── Post-commit side-effects ─────────────────────────────────
    # Run AFTER the transaction commits.  If email delivery fails the
    # booking is already safely in the database; we log but do not
    # surface the error — the user's booking succeeded.
    try:
        _send_booking_emails(booking_record)
    except Exception:
        logger.exception(
            "Failed to send booking notification emails for booking #%d.",
            booking_record.pk,
        )

    messages.success(
        request,
        "🎉 Your consultation has been booked! "
        "Our team will contact you within 30 minutes to confirm.",
    )
    return redirect("booking")   # PRG — final redirect on success


# ──────────────────────────────────────────────────────────────
# Blog views
# ──────────────────────────────────────────────────────────────

def blog(request):
    blogs_qs = Blog.objects.filter(status=True).order_by("-created_at")

    query = request.GET.get("search")
    if query:
        blogs_qs = blogs_qs.filter(
            Q(title__icontains=query) |
            Q(short_description__icontains=query) |
            Q(content__icontains=query)
        )

    category = request.GET.get("category")
    if category:
        blogs_qs = blogs_qs.filter(category__slug=category)

    paginator = Paginator(blogs_qs, 6)
    page = request.GET.get("page")
    blogs_page = paginator.get_page(page)

    # Sidebar data — fetched with select_related() to avoid N+1 on author/category
    featured_blog  = Blog.objects.filter(status=True, featured=True).select_related('category').first()
    latest_posts   = Blog.objects.filter(status=True).order_by("-created_at").select_related('category')[:5]
    popular_posts  = Blog.objects.filter(status=True).order_by("-views").select_related('category')[:5]
    top_blogs      = Blog.objects.filter(status=True).order_by("-views").select_related('category')[:3]
    categories     = Category.objects.all()

    context = {
        "blogs":             blogs_page,
        "featured_blog":     featured_blog,
        "latest_posts":      latest_posts,
        "popular_posts":     popular_posts,
        "top_blogs":         top_blogs,
        "categories":        categories,
        # FIXED: removed two extra DB queries (count + aggregate).
        # total_blogs now uses paginator.count which is already computed.
        "total_blogs":       paginator.count,
        "total_categories":  categories.count(),
    }

    return render(request, "home/blog.html", context)


def blog_detail(request, slug):
    """
    Render a single blog post.

    View counter
    ------------
    FIXED: was `blog.views += 1; blog.save()` which:
      1. Has a race condition under concurrent traffic (two requests both read
         views=100, both compute 101, both write 101 — actual should be 102).
      2. Calls full blog.save() which triggers auto_now on updated_at, making
         every page view appear to 'edit' the post.

    Fix: Use an atomic SQL UPDATE via F() expression — one database round-trip,
    no race condition, does NOT touch updated_at.
    """
    from django.db.models import F   # local import keeps top-level imports clean

    blog = get_object_or_404(
        Blog.objects.select_related('author', 'category'),
        slug=slug,
        status=True,      # prevent direct URL access to unpublished drafts
    )

    # Atomic increment — single UPDATE, no race condition
    Blog.objects.filter(pk=blog.pk).update(views=F('views') + 1)
    blog.views += 1   # update local object so template shows correct count

    latest_posts = Blog.objects.filter(status=True).order_by("-created_at")[:5]

    previous_blog = (
        Blog.objects
        .filter(status=True, created_at__gt=blog.created_at)
        .order_by("created_at")
        .first()
    )

    next_blog = (
        Blog.objects
        .filter(status=True, created_at__lt=blog.created_at)
        .order_by("-created_at")
        .first()
    )

    related = (
        Blog.objects
        .filter(category=blog.category, status=True)
        .exclude(id=blog.id)
        .select_related('category')[:3]
    )

    context = {
        "blog":          blog,
        "latest":        latest_posts,
        "previous_blog": previous_blog,
        "next_blog":     next_blog,
        "related":       related,
        "categories":    Category.objects.all(),
    }

    return render(request, "blog/blog_detail.html", context)
